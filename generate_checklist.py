import sys
import json
import uuid
import os
from openpyxl import load_workbook

def generate_json_from_excel(file_path, output_path="owasp_checklist.json"):
    wb = load_workbook(filename=file_path, data_only=True)
    sheet_names = wb.sheetnames[:12]  # Feuilles 1 à 12

    folders = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue

        section_title = rows[0][0] if rows[0][0] else f"Section {sheet_name}"
        folder = {
            "id": str(uuid.uuid4()),
            "title": str(section_title).strip(),
            "checklist": []
        }

        headers = None
        for row in rows:
            if row and "Test Name" in row:
                headers = {name: i for i, name in enumerate(row) if name}
                break

        if not headers or "Test Name" not in headers or "Status" not in headers:
            continue

        for row in rows:
            try:
                status = str(row[headers["Status"]]).strip().lower()
            except:
                continue
            if status != "not started":
                continue

            title = str(row[headers["Test Name"]]).strip()

            def get_col(name):
                idx = headers.get(name)
                return str(row[idx]).strip().replace("\n", "<br>") if idx is not None and row[idx] else ""

            wstg_id = get_col("WSTG-ID")
            objectives = get_col("Objectives") or get_col("Description")
            tools = get_col("Tools")
            owasp_top10 = get_col("OWASP Top 10")

            content_html = f"""
                <p><strong>WSTG-ID:</strong> {wstg_id} <br><br> </p>
                <p><strong>Objectives:</strong> {objectives} <br> <br></p>
                <p><strong>Tools:</strong> {tools} <br><br> </p>
                <p><strong>OWASP Top 10:</strong> {owasp_top10}<br><br></p>
                <p><strong> Result :</strong> <br><br> </p>
            """.strip()

            check = {
                "id": str(uuid.uuid4()),
                "title": title,
                "content": content_html
            }
            folder["checklist"].append(check)

        folders.append(folder)

    final_data = {
        "targets": [],
        "libraries": [{
            "id": str(uuid.uuid4()),
            "title": "OWASP Testing Checklist",
            "folders": folders
        }],
        "templates": [],
        "payloads": [],
        "messages": {
            "showDeleteConfirmation": True
        }
    }

    # Enregistrer dans un fichier JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(final_data, f, indent=2, ensure_ascii=False)

    print(f"✅ JSON généré et enregistré dans : {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python script.py <excel-file> [output-json]")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) >= 3 else "owasp_checklist.json"

    generate_json_from_excel(excel_file, output_file)
